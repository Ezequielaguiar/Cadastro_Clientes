import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import pathlib
import openpyxl , xlrd
import openpyxl 

#criando tema padrão do Custumer tkinter
ctk.set_appearance_mode('Dark')
ctk.set_default_color_theme('blue')

class configurar_janela(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.configurar_tela()
        self.tela()
        self.cabecalho()
        self.titulo_input()
        self.combobox()
        self.variaveis()
        self.input_usuario()
        self.botoes()
        
    def configurar_tela(self):
        self.title( 'Sistema Cadastro de Clientes')
        self.geometry('700x500') 
    
    def alterar_tema(self,novo_tema):
        ctk.set_appearance_mode(novo_tema)    
    
    def tela(self):
        self.tipo_tema = ctk.CTkLabel(self,text='Tema',bg_color='transparent',text_color=['#000','#fff'])
        self.tipo_tema.place(x=50,y=430)
        self.menu = ctk.CTkOptionMenu(self, values=['Dark','Light'],command=self.alterar_tema)
        self.menu.place(x=50,y=460)
      
    def cabecalho(self):
        faixa = ctk.CTkFrame(self,width = 700 , height= 50,corner_radius=0,bg_color=  'teal', fg_color='teal')
        faixa.place(x=0,y=10)
        sub_titulo = ctk.CTkLabel(faixa,text='Sistema de cadastro de clientes',font=('Century Gothic bold',24),text_color='#fff')
        sub_titulo.place(x=0,y=10)
        instrucao = ctk.CTkLabel(self,text='Por favor, preencha todos os campos do formulario!',font=('Century Gothic bold',16),text_color=['#000','#fff'])
        instrucao.place(x=50,y=70)
    
    def variaveis(self):
        self.valor_nome = StringVar()
        self.valor_contato = StringVar()
        self.valor_idade = StringVar()
        self.valor_endereco = StringVar()

           
    def input_usuario(self):
        nome = ctk.CTkEntry(self,width=350,textvariable=self.valor_nome,font=('Century Gohtic',16),fg_color='transparent')
        nome.place(x=50,y= 150)
        contato = ctk.CTkEntry(self,width=200,textvariable=self.valor_contato, font = ('Century Gohtic',16),fg_color = 'transparent')
        contato.place(x=450,y=150)
        idade = ctk.CTkEntry(self,width=150,textvariable=self.valor_idade, font=('Century Gohtic',16),fg_color='transparent')
        idade.place(x=300,y=220)
        endereco = ctk.CTkEntry(self,width=200,textvariable=self.valor_endereco , font=('Century Gohtic',16),fg_color='transparent')
        endereco.place(x=50,y=220)
        self.observacoes = ctk.CTkTextbox(self,width=500,height=150,font=('arial',18),border_color='#aaa',border_width=2,fg_color='transparent')
        self.observacoes.place(x = 180,y=260)
        
        
    def combobox(self):
        sexo = ctk.CTkComboBox(self,values=['Masculino','Feminino'],font=('Century Gothic bold',14))
        sexo.set('Sexo')
        sexo.place(x=500,y=220)
    
    def titulo_input(self):
        titulo_nome = ctk.CTkLabel(self,text='Nome completo',font=('Century Gothic bold',16),text_color=['#000','#fff'])    
        titulo_nome.place(x=50,y=120)
        titulo_contato = ctk.CTkLabel(self,text= 'Contato',font=('Century Gothic bold',16),text_color=['#000','#fff'])
        titulo_contato.place(x=450,y=120)
        titulo_idade = ctk.CTkLabel(self,text='Idade',font=('Century Gothic bold',16),text_color=['#000','#fff'])
        titulo_idade.place(x=300,y=190)
        titulo_endereco = ctk.CTkLabel(self,text='Endereço',font=('Century Gothic bold',16),text_color=['#000','#fff'])
        titulo_endereco.place(x=50,y=190)
        titulo_sexo = ctk.CTkLabel(self,text='Genero',font=('Century Gothic bold',16),text_color=['#000','#fff'])
        titulo_sexo.place(x = 500,y=190)
        titulo_obs = ctk.CTkLabel(self,text='Observações',font=('Century Gothic bold',16),text_color=['#000','#fff'])
        titulo_obs.place(x=50,y=260)
    
    def salvar(self):
        self.name = self.valor_nome.get()
        self.contato = self.valor_contato.get()
        self.idade = self.valor_idade.get()
        self.endereco = self.valor_endereco.get()
        self.obs = self.observacoes.get(0.0,END)
        if (self.name == '' or self.contato == '' or  self.idade =='' or self.endereco=='' ):
            messagebox.showerror('Sistema','ERRO!\nPor favor preecher todos os campos!')
        else:   
            self.obter_dados()
        
    
    def clear(self):
        self.valor_nome.set('')   
        self.valor_contato.set('')
        self.valor_idade.set('')
        self.valor_endereco.set('')
        self.observacoes.delete(0.0,END)
        
    def obter_dados(self):

        ficheiro = pathlib.Path('Cadastro.xlsx')        
        
        if ficheiro.exists():
            pass
        else:
            ficheiro = openpyxl.Workbook()
            cabecalho_excel = ficheiro.active
            cabecalho_excel['A1'] = 'NOME COMPLETO'
            cabecalho_excel['B1'] = 'CONTATO'
            cabecalho_excel['C1'] = 'IDADE'
            cabecalho_excel['D1'] = 'GENERO'
            cabecalho_excel['E1'] = 'ENDEREÇO' 
            ficheiro.save('Cadastro.xlsx')
            
        ficheiro = openpyxl.load_workbook('Cadastro.xlsx')
        celula = ficheiro.active
        celula.cell(column= 1,row=celula.max_row+1,value = self.name)
        celula.cell(column= 2,row=celula.max_row,value = self.contato)
        celula.cell(column= 3,row=celula.max_row,value = self.idade)
        celula.cell(column= 4,row=celula.max_row,value = self.endereco)
        celula.cell(column= 5,row=celula.max_row,value = self.obs)
        ficheiro.save(r'Cadastro.xlsx')
        messagebox.showinfo('Sisyema','Dados salvos com sucesso!')
        
    def botoes(self):
        gravar = ctk.CTkButton(self,text = 'Salvar'.upper(),command= self.salvar,fg_color='#151',hover_color='#131') 
        gravar.place(x = 300,y=420)
        limpar = ctk.CTkButton(self,text='Limpar'.upper(),command=self.clear,fg_color='#555')
        limpar.place(x=500,y=420)
        
if __name__ == "__main__":
    janela = configurar_janela()
    janela.mainloop()
        